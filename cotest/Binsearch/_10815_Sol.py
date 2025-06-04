import sys
print("--- DEBUG: mkexcel.py executing from:", __file__)
print("--- DEBUG: sys.path:", sys.path)
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
# from openpyxl.chart.axis import ChartAxis # 이 줄이 완전히 삭제되었습니다.
from openpyxl.styles import Font, Alignment # 이 줄은 사용되지 않지만, 다른 스타일링에 필요할 수 있어 일단 유지합니다.

def create_target_salary_simulation_excel(
    data_file_path='data.xlsx', # 실수령액 데이터 파일 경로
    start_year=2026,
    hong_gil_dong_annual_salary=3000, # 만원
    hong_gil_dong_avg_asset_growth_rate=0.20, # 20%
    hong_gil_dong_zero_profit_cycle=5, # 5년에 한번 0% 수익
    hong_gil_dong_start_month=9, # 홍길동 소득 시작 월 (9월부터)
    user_initial_annual_salary=3500, # 만원 (시작 연봉)
    user_annual_salary_increase_rate=0.03, # 사용자 연간 연봉 상승률 (예: 3%)
    user_job_change_cycle=3, # 사용자 이직 주기 (년)
    user_job_change_salary_increase_rate=0.08, # 이직 시점 추가 연봉 상승률 (예: 8%)
    user_monthly_housing_cost=400000, # 원 (1인 가구 평균 예시: 40만원)
    user_monthly_fixed_living_cost=1000000, # 원 (1인 가구 평균 예시: 100만원)
    user_annual_investment_return_rate=0.08, # 사용자 연간 투자 수익률 (예: 8%)
    simulation_years=30 # 시뮬레이션 기간
):
    """
    주어진 데이터를 바탕으로 홍길동보다 높은 자산과 소득을 가지기 위해
    사용자님의 연봉이 연도별로 얼마나 높아야 하는지 시뮬레이션 엑셀 파일을 생성합니다.
    실수령액 데이터는 data.xlsx 파일에서 불러옵니다.
    생성된 엑셀 파일 내부에 누계 자산 그래프와 소득 그래프를 삽입합니다.
    """

    # ----------------------------------------------------
    # 1. 실수령액 및 공제액 테이블 데이터 로드
    # ----------------------------------------------------
    try:
        salary_df = pd.read_excel(data_file_path, sheet_name='Sheet1')
        
        # '연봉' 열이 있는지 확인
        if '연봉' not in salary_df.columns:
            raise ValueError(f"'{data_file_path}'의 'Sheet1'에 '연봉' 열이 없습니다. 열 이름을 확인해주세요.")

        # 필수 열 확인 (여기서는 '연봉'으로 체크)
        required_columns = ['연봉', '실수령액', '공제액계', '국민연금', '건강보험', '장기요양', '고용보험', '소득세', '지방소득세']
        if not all(col in salary_df.columns for col in required_columns):
            raise ValueError(f"'{data_file_path}'의 'Sheet1'에 다음 필수 열 중 일부가 누락되었습니다: {required_columns}")

        # '연봉' 열을 숫자형으로 변환하는 함수 (내부 계산용)
        def clean_salary_col_to_int(s):
            if isinstance(s, str):
                return int(s.replace(',', '').replace('만원', '').strip())
            # 이미 숫자형일 경우 (예: Excel에서 숫자로 입력된 경우)
            return int(s)

        # '연봉' 열을 숫자로 변환하여 새로운 열 '연봉_숫자' 생성
        salary_df['연봉_숫자'] = salary_df['연봉'].apply(clean_salary_col_to_int)

    except FileNotFoundError:
        print(f"오류: '{data_file_path}' 파일을 찾을 수 없습니다. 스크립트와 같은 폴더에 두셨는지 확인해주세요.")
        return
    except Exception as e:
        print(f"오류: '{data_file_path}' 파일을 읽는 중 문제가 발생했습니다: {e}")
        return

    def get_salary_info(annual_salary_in_krw_10k):
        """
        연봉(만원)에 해당하는 실수령액 및 공제액 정보를 반환합니다.
        가장 가까운 연봉(만원)을 찾아서 정보를 가져옵니다.
        '연봉_숫자' 열을 사용하여 계산합니다.
        """
        # 연봉 테이블이 100만원 단위로 되어 있으므로, 입력된 연봉을 100만원 단위로 반올림하여 매칭
        lookup_salary = round(annual_salary_in_krw_10k / 100) * 100
        
        # 테이블 범위 내에 있는지 확인 (이제 '연봉_숫자' 열 사용)
        if lookup_salary < salary_df['연봉_숫자'].min():
            lookup_salary = salary_df['연봉_숫자'].min()
        elif lookup_salary > salary_df['연봉_숫자'].max():
            lookup_salary = salary_df['연봉_숫자'].max() # 테이블 최대치로 제한

        # 가장 가까운 연봉을 찾아 해당 행의 데이터를 반환 (이제 '연봉_숫자' 열 사용)
        closest_row_index = (salary_df['연봉_숫자'] - lookup_salary).abs().idxmin()
        return salary_df.loc[closest_row_index]

    # ----------------------------------------------------
    # 2. 시뮬레이션 데이터프레임 초기화
    # ----------------------------------------------------
    years = list(range(start_year, start_year + simulation_years))
    df = pd.DataFrame({'연도': years})

    # ----------------------------------------------------
    # 3. 홍길동 자산 시뮬레이션 (고정)
    # ----------------------------------------------------
    hong_gil_dong_annual_salary_krw_10k = hong_gil_dong_annual_salary
    hong_gil_dong_info = get_salary_info(hong_gil_dong_annual_salary_krw_10k)
    hong_gil_dong_monthly_net_salary = hong_gil_dong_info['실수령액']

    hong_gil_dong_asset_end_of_year = 0
    hong_gil_dong_assets = []
    hong_gil_dong_annual_returns_calculated = []
    hong_gil_dong_annual_savings_list = []
    hong_gil_dong_actual_annual_salary_list = [] # 첫 해 부분 소득 반영 (만원)

    for i, year in enumerate(years):
        current_asset = hong_gil_dong_asset_end_of_year
        
        annual_growth_rate = hong_gil_dong_avg_asset_growth_rate
        if (i + 1) % hong_gil_dong_zero_profit_cycle == 0:
            annual_growth_rate = 0 # 0% 수익 연도

        # 홍길동의 연간 저축 가능액 계산 (첫 해는 9월부터, 즉 4개월치만 반영)
        months_in_year = 12
        if i == 0: # 첫 해 (start_year)
            months_in_year = 12 - hong_gil_dong_start_month + 1 # 9월부터 시작 시 4개월 (9, 10, 11, 12)
        
        # 순소득 정의: 실수령액 - (주거비 + 고정생활비)
        hong_gil_dong_monthly_net_income = hong_gil_dong_monthly_net_salary - (user_monthly_housing_cost + user_monthly_fixed_living_cost)
        hong_gil_dong_annual_investment = max(0, hong_gil_dong_monthly_net_income * months_in_year)
        
        annual_return = current_asset * annual_growth_rate
        hong_gil_dong_asset_end_of_year = current_asset + hong_gil_dong_annual_investment + annual_return
        
        hong_gil_dong_assets.append(hong_gil_dong_asset_end_of_year)
        hong_gil_dong_annual_returns_calculated.append(annual_return)
        hong_gil_dong_annual_savings_list.append(hong_gil_dong_annual_investment)
        # 첫 해는 월 할 연봉으로 계산하여 표시
        hong_gil_dong_actual_annual_salary_list.append(hong_gil_dong_annual_salary * (months_in_year / 12))


    df['홍길동 연봉(만원)'] = hong_gil_dong_actual_annual_salary_list # 첫 해 부분 연봉 반영
    df['홍길동 월 실수령액(원)'] = hong_gil_dong_monthly_net_salary
    df['홍길동 연간 저축 가능액(원)'] = hong_gil_dong_annual_savings_list
    df['홍길동 연간 자산 증식률'] = [hong_gil_dong_avg_asset_growth_rate if (i + 1) % hong_gil_dong_zero_profit_cycle != 0 else 0 for i in range(simulation_years)]
    df['홍길동 연간 투자 수익(원)'] = hong_gil_dong_annual_returns_calculated
    df['홍길동 누적 자산(원)'] = hong_gil_dong_assets

    # ----------------------------------------------------
    # 4. 사용자 자산 시뮬레이션
    # ----------------------------------------------------
    user_yearly_data = []
    user_current_asset = 0
    current_user_annual_salary_krw_10k = user_initial_annual_salary

    for i, year in enumerate(years):
        # 연봉 상승 로직 적용 (매년 기본 상승 + 이직 주기마다 추가 상승)
        if i > 0:
            current_user_annual_salary_krw_10k *= (1 + user_annual_salary_increase_rate)
            if i % user_job_change_cycle == 0: # 이직 주기에 맞춰 추가 상승
                current_user_annual_salary_krw_10k *= (1 + user_job_change_salary_increase_rate)
        
        # 연봉 테이블 범위에 맞게 조정 (반올림 전)
        current_user_annual_salary_krw_10k = max(salary_df['연봉_숫자'].min(), min(salary_df['연봉_숫자'].max(), current_user_annual_salary_krw_10k))

        # 실제 사용될 연봉(만원)은 반올림된 값으로 get_salary_info에 전달
        user_info = get_salary_info(current_user_annual_salary_krw_10k)
        
        user_monthly_net_salary = user_info['실수령액']
        user_monthly_deductions_total = user_info['공제액계']
        user_national_pension = user_info['국민연금']
        user_health_insurance = user_info['건강보험']
        user_longterm_care_insurance = user_info['장기요양']
        user_employment_insurance = user_info['고용보험']
        user_income_tax = user_info['소득세']
        user_local_income_tax = user_info['지방소득세']

        # 순소득 정의: 실수령액 - (주거비 + 고정생활비)
        user_monthly_net_income = user_monthly_net_salary - (user_monthly_housing_cost + user_monthly_fixed_living_cost)
        user_annual_investment = max(0, user_monthly_net_income * 12) # 순소득이 음수면 저축 가능액은 0

        user_annual_return_from_asset = user_current_asset * user_annual_investment_return_rate # 사용자 투자 수익률 적용
        user_current_asset = user_current_asset + user_annual_investment + user_annual_return_from_asset
        
        user_yearly_data.append({
            '연도': year,
            '사용자 연봉(만원)': round(current_user_annual_salary_krw_10k), # 실제 사용된 연봉(만원)은 반올림
            '사용자 월 실수령액(원)': user_monthly_net_salary,
            '사용자 월 공제액 합계(원)': user_monthly_deductions_total,
            '사용자 월 국민연금(원)': user_national_pension,
            '사용자 월 건강보험(원)': user_health_insurance,
            '사용자 월 장기요양(원)': user_longterm_care_insurance,
            '사용자 월 고용보험(원)': user_employment_insurance,
            '사용자 월 소득세(원)': user_income_tax,
            '사용자 월 지방소득세(원)': user_local_income_tax,
            '사용자 월 주거비(원)': user_monthly_housing_cost,
            '사용자 월 고정생활비(원)': user_monthly_fixed_living_cost,
            '사용자 월 순소득(원)': user_monthly_net_income,
            '사용자 연간 저축 가능액(원)': user_annual_investment,
            '사용자 연간 투자 수익률': user_annual_investment_return_rate,
            '사용자 연간 투자 수익(원)': user_annual_return_from_asset,
            '사용자 누적 자산(원)': user_current_asset
        })

    user_df = pd.DataFrame(user_yearly_data)
    
    # ----------------------------------------------------
    # 5. 최종 데이터프레임 병합 및 정리
    # ----------------------------------------------------
    final_df = pd.merge(df, user_df, on='연도', how='left')

    # 홍길동과 사용자 자산 차이 계산
    final_df['자산 증식 속도 차이(원) (홍길동 - 사용자)'] = final_df['홍길동 누적 자산(원)'] - final_df['사용자 누적 자산(원)']

    # ----------------------------------------------------
    # 6. 엑셀 파일로 저장 및 차트 삽입
    # ----------------------------------------------------
    current_time_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_name = f"나의_목표_연봉_시뮬레이션_{current_time_str}.xlsx"
    
    # 새로운 Workbook 생성
    wb = Workbook()

    # 기본 설정 시트
    ws_settings = wb.active
    ws_settings.title = '기본 설정'
    
    settings_data = {
        '항목': [
            '시뮬레이션 시작 연도',
            '시뮬레이션 기간(년)',
            '홍길동 연봉(만원)',
            '홍길동 연평균 자산 증식률',
            '홍길동 자산 증식률 0% 주기(년)',
            '홍길동 소득 시작 월 (월 할 계산)',
            '사용자 시작 연봉(만원)',
            '사용자 연간 기본 연봉 상승률',
            '사용자 이직 주기(년)',
            '사용자 이직 시점 추가 연봉 상승률',
            '사용자 월 주거비(원)',
            '사용자 월 고정생활비(원)',
            '사용자 연간 투자 수익률'
        ],
        '값': [
            start_year,
            simulation_years,
            hong_gil_dong_annual_salary,
            f"{hong_gil_dong_avg_asset_growth_rate*100:.0f}%",
            hong_gil_dong_zero_profit_cycle,
            hong_gil_dong_start_month,
            user_initial_annual_salary,
            f"{user_annual_salary_increase_rate*100:.1f}%",
            user_job_change_cycle,
            f"{user_job_change_salary_increase_rate*100:.1f}%",
            f"{user_monthly_housing_cost:,.0f}",
            f"{user_monthly_fixed_living_cost:,.0f}",
            f"{user_annual_investment_return_rate*100:.1f}%"
        ]
    }
    settings_df = pd.DataFrame(settings_data)
    for r_idx, row in enumerate(dataframe_to_rows(settings_df, index=False, header=True), 1):
        ws_settings.append(row)
    # 컬럼 너비 자동 조정
    for col in ws_settings.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if cell.value is not None: # None 값 처리
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_settings.column_dimensions[column].width = adjusted_width

    # 사용자 연봉 계산 시트 (상세 내역)
    ws_user_salary = wb.create_sheet('사용자 연봉_세부')
    user_salary_deduction_df = final_df[[
        '연도',
        '사용자 연봉(만원)',
        '사용자 월 실수령액(원)',
        '사용자 월 공제액 합계(원)',
        '사용자 월 국민연금(원)',
        '사용자 월 건강보험(원)',
        '사용자 월 장기요양(원)',
        '사용자 월 고용보험(원)',
        '사용자 월 소득세(원)',
        '사용자 월 지방소득세(원)',
        '사용자 월 주거비(원)',
        '사용자 월 고정생활비(원)',
        '사용자 월 순소득(원)',
        '사용자 연간 저축 가능액(원)'
    ]]
    for r_idx, row in enumerate(dataframe_to_rows(user_salary_deduction_df, index=False, header=True), 1):
        ws_user_salary.append(row)
    # 컬럼 너비 자동 조정
    for col in ws_user_salary.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if cell.value is not None: # None 값 처리
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_user_salary.column_dimensions[column].width = adjusted_width

    # 자산 증식 시뮬레이션 시트
    ws_simulation = wb.create_sheet('자산_시뮬레이션')
    final_df_simulation = final_df[[
        '연도',
        '홍길동 연봉(만원)',
        '홍길동 연간 저축 가능액(원)',
        '홍길동 연간 투자 수익(원)',
        '홍길동 누적 자산(원)',
        '사용자 연봉(만원)',
        '사용자 연간 저축 가능액(원)',
        '사용자 연간 투자 수익(원)',
        '사용자 누적 자산(원)',
        '자산 증식 속도 차이(원) (홍길동 - 사용자)'
    ]]
    for r_idx, row in enumerate(dataframe_to_rows(final_df_simulation, index=False, header=True), 1):
        ws_simulation.append(row)
    # 컬럼 너비 자동 조정
    for col in ws_simulation.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if cell.value is not None: # None 값 처리
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_simulation.column_dimensions[column].width = adjusted_width


    # 세액 테이블 시트
    ws_salary_table = wb.create_sheet('세액_테이블')
    # '연봉_숫자' 열은 엑셀 파일 저장 시 포함하지 않음
    salary_df_for_excel = salary_df.drop(columns=['연봉_숫자'], errors='ignore')
    for r_idx, row in enumerate(dataframe_to_rows(salary_df_for_excel, index=False, header=True), 1):
        ws_salary_table.append(row)
    # 컬럼 너비 자동 조정
    for col in ws_salary_table.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if cell.value is not None: # None 값 처리
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_salary_table.column_dimensions[column].width = adjusted_width


    # ----------------------------------------------------
    # 7. 엑셀 시트 내부에 차트 삽입
    # ----------------------------------------------------

    # 엑셀 시트 행/열 인덱스 계산 (데이터프레임이 1부터 시작하므로 +1)
    # '자산_시뮬레이션' 시트에 데이터가 들어가므로 해당 시트를 기준으로 합니다.
    data_start_row = 2 # 헤더 다음 행부터 데이터 시작
    data_end_row = len(final_df_simulation) + 1 # 데이터 마지막 행 (헤더 포함)
    
    # 누적 자산 그래프
    chart1 = LineChart()
    chart1.title = "홍길동과 사용자 누적 자산 비교"
    chart1.style = 10 # 차트 스타일 (색상, 선 굵기 등)
    chart1.x_axis.title = "연도"
    chart1.y_axis.title = "누적 자산 (원)"
    # Y축 라벨을 숫자로만 표기 (Excel에서 자체적으로 화폐 단위 포맷팅 가능)
    chart1.y_axis.number_format = '#,##0' # 쉼표로 천 단위 구분

    # 데이터 참조 (Values)
    # 홍길동 누적 자산: E열 (5번째 컬럼, index 4)
    values1 = Reference(ws_simulation, min_col=5, min_row=data_start_row, max_col=5, max_row=data_end_row)
    # 사용자 누적 자산: I열 (9번째 컬럼, index 8)
    values2 = Reference(ws_simulation, min_col=9, min_row=data_start_row, max_col=9, max_row=data_end_row)
    
    # 카테고리 (X축 연도) - A열 (1번째 컬럼, index 0)
    years_ref = Reference(ws_simulation, min_col=1, min_row=data_start_row, max_col=1, max_row=data_end_row)
    chart1.set_categories(years_ref)

    # 시리즈 추가
    chart1.add_data(values1, titles_from_data=False)
    s1 = chart1.series[0]
    s1.title = "홍길동 누적 자산"
    
    chart1.add_data(values2, titles_from_data=False)
    s2 = chart1.series[1]
    s2.title = "사용자 누적 자산"

    # 차트 위치 지정 (예: 'K1' 셀부터 시작)
    ws_simulation.add_chart(chart1, "K1")

    # 소득 그래프
    chart2 = LineChart()
    chart2.title = "홍길동과 사용자 연봉 비교"
    chart2.style = 10
    chart2.x_axis.title = "연도"
    chart2.y_axis.title = "연봉 (만원)"
    chart2.y_axis.number_format = '#,##0"만원"' # '만원' 단위 추가

    # 데이터 참조 (Values)
    # 홍길동 연봉(만원): B열 (2번째 컬럼, index 1)
    values3 = Reference(ws_simulation, min_col=2, min_row=data_start_row, max_col=2, max_row=data_end_row)
    # 사용자 연봉(만원): F열 (6번째 컬럼, index 5)
    values4 = Reference(ws_simulation, min_col=6, min_row=data_start_row, max_col=6, max_row=data_end_row)

    # 카테고리 (X축 연도) - A열 (1번째 컬럼, index 0)
    chart2.set_categories(years_ref) # 동일한 연도 참조 사용

    # 시리즈 추가
    chart2.add_data(values3, titles_from_data=False)
    s3 = chart2.series[0]
    s3.title = "홍길동 연봉"
    
    chart2.add_data(values4, titles_from_data=False)
    s4 = chart2.series[1]
    s4.title = "사용자 연봉"

    # 차트 위치 지정 (예: 'K20' 셀부터 시작)
    ws_simulation.add_chart(chart2, "K20")
    
    # 워크북 저장
    wb.save(file_name)

    print(f"'{file_name}' 파일이 성공적으로 생성되었습니다.")
    print(f"\n[중요] 시뮬레이션에 사용된 실수령액 데이터는 '{data_file_path}' 파일의 'Sheet1'에서 불러왔습니다.")
    print("      이 파일이 스크립트와 같은 폴더에 있어야 합니다.")
    print("      엑셀 파일을 열어 '기본 설정' 시트에서 다양한 변수들을 조정해 보세요!")
    print("      차트는 '자산_시뮬레이션' 시트에 삽입되었습니다.")


# 함수 실행 (원하는 값으로 파라미터를 변경하여 실행할 수 있습니다.)
create_target_salary_simulation_excel(
    data_file_path='data.xlsx', # 여기에 data.xlsx 파일 경로를 넣어주세요.
    user_initial_annual_salary=3500, # 시작 연봉 3500만원
    user_annual_salary_increase_rate=0.03, # 사용자 연간 기본 연봉 상승률 3%
    user_job_change_cycle=3, # 사용자 이직 주기 3년
    user_job_change_salary_increase_rate=0.08, # 이직 시점 추가 연봉 상승률 8%
    user_monthly_housing_cost=400000, # 월 주거비 40만원
    user_monthly_fixed_living_cost=1000000, # 월 고정생활비 100만원
    user_annual_investment_return_rate=0.08, # 사용자 연간 투자 수익률 8% (이 값을 높이면 더 낮은 연봉으로도 따라잡기 가능)
    simulation_years=30 # 30년간 시뮬레이션
)