import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_financial_simulation_excel(start_year=2025,
                                      hong_gil_dong_initial_asset=40_000_000,
                                      user_initial_asset=0,
                                      initial_user_housing_cost_monthly=500_000, # 초기값
                                      initial_user_fixed_living_cost_monthly=1_000_000): # 초기값

    # --- 설정 변수 (사용자가 엑셀에서 수정 가능하도록) ---
    settings = {
        '홍길동 연평균 자산 증식률 (%)': 20.0,
        '홍길동 5년 주기 0% 수익 여부 (True/False)': True,
        '사용자 연봉 시작 (만원)': 3500.0,
        '사용자 기본 연봉 상승률 (%)': 3.0,
        '사용자 이직 주기 (년)': 3,
        '사용자 이직 시 연봉 상승률 (%)': 8.0,
        '사용자 연평균 자산 증식률 (%)': 7.0, # 초기 사용자 투자 수익률 (조절 가능)
        '월 주거비 (원)': initial_user_housing_cost_monthly,
        '월 고정 생활비 (원)': initial_user_fixed_living_cost_monthly,
        '시뮬레이션 기간 (년)': 30 # 시뮬레이션할 기간
    }

    # --- 엑셀 파일 생성 ---
    file_name = '재테크_시뮬레이션.xlsx'
    writer = pd.ExcelWriter(file_name, engine='openpyxl')

    # --- Sheet1: 실수령액 테이블 (사용자 입력용) ---
    # 실제 실수령액 데이터를 여기에 채워 넣으세요.
    # 예시 데이터 (실제 값과 다를 수 있음)
    salary_data_example = {
        '연봉': [1000, 1100, 1200, 1300, 1400, 1500, 2000, 2500, 3000, 3500, 4000, 4500, 5000, 6000, 7000, 8000, 9000, 10000, 11000, 12000, 13000, 14000, 15000, 20000, 25000, 30000, 35000, 40000, 45000, 50000, 55000, 60000, 65000, 70000, 75000, 80000, 85000, 90000, 95000, 100000],
        '실수령액': [771033, 847286, 923500, 998863, 1073736, 1148580, 1528000, 1900000, 2270000, 2630000, 2980000, 3320000, 3660000, 4320000, 4980000, 5640000, 6300000, 6960000, 7500000, 8000000, 8500000, 9000000, 9500000, 12000000, 14500000, 17000000, 19500000, 22000000, 24500000, 27000000, 29500000, 32000000, 34500000, 37000000, 39500000, 42000000, 44500000, 47000000, 49500000, 52000000] # 임의의 값, 실제 데이터로 대체 필요
    }
    df_salary_lookup = pd.DataFrame(salary_data_example)
    df_salary_lookup['연봉'] = df_salary_lookup['연봉'] * 10000 # 만원 단위를 원 단위로 변경

    # Sheet1에 쓰기
    df_salary_lookup.to_excel(writer, sheet_name='실수령액 테이블', index=False)
    # 실수령액 테이블 시트 꾸미기 (추후 사용자가 여기에 실제 데이터를 입력하도록 가이드)
    ws_salary_table = writer.sheets['실수령액 테이블']
    ws_salary_table['A1'].value = '연봉 (원)'
    ws_salary_table['B1'].value = '실수령액 (월, 원)'
    ws_salary_table.freeze_panes = 'A2' # 첫 행 고정
    for col_idx in range(1, len(df_salary_lookup.columns) + 1):
        ws_salary_table.cell(row=1, column=col_idx).font = Font(bold=True)
        ws_salary_table.column_dimensions[get_column_letter(col_idx)].width = 15
    ws_salary_table.append(['', ''])
    ws_salary_table.append(['**여기에 실제 연봉별 월 실수령액 데이터를 입력해주세요.**', ''])
    ws_salary_table.append(['(연봉은 원 단위, 실수령액은 월 단위 원)'])


    # --- Sheet2: 시뮬레이션 ---
    data = []
    hong_gil_dong_asset = hong_gil_dong_initial_asset
    user_asset = user_initial_asset
    user_current_annual_salary = settings['사용자 연봉 시작 (만원)'] * 10000 # 원 단위로 변환

    hong_gil_dong_passed_zero_profit_years = 0 # 0% 수익이 지난 몇 년 동안 지속되었는지 추적
    user_reach_hong_gil_dong_asset_year = None # 사용자가 홍길동 자산을 따라잡는 연도

    # 시뮬레이션 루프
    for year_idx in range(settings['시뮬레이션 기간 (년)']):
        current_year = start_year + year_idx
        
        # --- 홍길동 소득 및 자산 계산 ---
        # 2026년 9월부터 연봉 3천만원 시작 (첫 해는 4개월치 소득 반영)
        hong_gil_dong_annual_salary_this_year = 0
        hong_gil_dong_monthly_net_income = 0
        if current_year >= 2026:
            hong_gil_dong_annual_salary_this_year = 30_000_000 # 연 3천만원
            
            # 실수령액 조회 (가장 근접한 연봉을 찾음)
            # 홍길동은 3천만원 고정으로 가정, 해당 연봉의 실수령액을 찾아야 함
            # 여기서는 편의상 연봉 3천만원의 월 실수령액이 227만원이라고 가정 (위 예시 데이터에서 가져옴)
            # 실제로는 Sheet1에서 정확히 찾아야 함
            if 30_000_000 in df_salary_lookup['연봉'].values:
                hong_gil_dong_monthly_take_home = df_salary_lookup[df_salary_lookup['연봉'] == 30_000_000]['실수령액'].iloc[0]
            else:
                # 가장 가까운 연봉의 실수령액을 찾거나, 직접 계산하는 로직 추가 필요 (지금은 예시값)
                hong_gil_dong_monthly_take_home = 2_270_000 # 임의의 값

            # 2026년 첫 해는 9월부터 (4개월)
            if current_year == 2026:
                hong_gil_dong_monthly_net_income = hong_gil_dong_monthly_take_home # 주거비, 생활비는 따로 계산하지 않음 (자산 증식률에 포함된 것으로 가정)
            else:
                hong_gil_dong_monthly_net_income = hong_gil_dong_monthly_take_home

        hong_gil_dong_investment_amount = hong_gil_dong_monthly_net_income * 12 # 연간 투자 가능 금액

        # 홍길동 자산 증식
        hong_gil_dong_asset_growth_rate = settings['홍길동 연평균 자산 증식률 (%)'] / 100
        
        if settings['홍길동 5년 주기 0% 수익 여부 (True/False)'] and (year_idx + 1) % 5 == 0:
            hong_gil_dong_annual_asset_growth_rate = 0 # 5년에 한번 0% 수익
            hong_gil_dong_passed_zero_profit_years = 0 # 0% 수익 발생했으므로 초기화
        else:
            hong_gil_dong_annual_asset_growth_rate = hong_gil_dong_asset_growth_rate
            hong_gil_dong_passed_zero_profit_years += 1 # 0% 수익이 아닌 해는 카운트 증가

        hong_gil_dong_asset = (hong_gil_dong_asset + hong_gil_dong_investment_amount) * (1 + hong_gil_dong_annual_asset_growth_rate)
        
        # --- 사용자 소득 및 자산 계산 ---
        # 연봉 상승
        if year_idx > 0: # 첫 해 이후부터 적용
            if year_idx % settings['사용자 이직 주기 (년)'] == 0:
                user_current_annual_salary *= (1 + settings['사용자 이직 시 연봉 상승률 (%)'] / 100)
            else:
                user_current_annual_salary *= (1 + settings['사용자 기본 연봉 상승률 (%)'] / 100)
        
        # 실수령액 조회 (가장 근접한 연봉을 찾음)
        # Sheet1의 연봉 데이터를 활용해야 함
        # 여기서는 가장 가까운 연봉을 찾아 실수령액을 가져오는 로직 (보간법 X, 단순 근접)
        # 연봉 데이터가 정렬되어 있다고 가정
        closest_salary_row = df_salary_lookup.iloc[(df_salary_lookup['연봉'] - user_current_annual_salary).abs().argsort()[:1]]
        user_monthly_take_home = closest_salary_row['실수령액'].iloc[0] if not closest_salary_row.empty else 0
        
        user_annual_take_home = user_monthly_take_home * 12
        
        # 순소득 계산 (절대값)
        user_annual_housing_cost = settings['월 주거비 (원)'] * 12
        user_annual_fixed_living_cost = settings['월 고정 생활비 (원)'] * 12
        
        # 순소득: 실수령액 - (주거비 + 고정생활비)
        user_net_income = user_annual_take_home - (user_annual_housing_cost + user_annual_fixed_living_cost)
        
        # 투자 금액 (순소득이 음수이면 0)
        user_investment_amount = max(0, user_net_income)

        # 사용자 자산 증식
        user_asset_growth_rate = settings['사용자 연평균 자산 증식률 (%)'] / 100
        user_asset = (user_asset + user_investment_amount) * (1 + user_asset_growth_rate)

        # 사용자가 홍길동 자산을 따라잡는 시점 기록
        if user_reach_hong_gil_dong_asset_year is None and user_asset >= hong_gil_dong_asset:
            user_reach_hong_gil_dong_asset_year = current_year
        
        data.append({
            '연도': current_year,
            '홍길동_연봉 (원)': hong_gil_dong_annual_salary_this_year,
            '홍길동_연간_투자금액 (원)': hong_gil_dong_investment_amount,
            '홍길동_연말_자산 (원)': hong_gil_dong_asset,
            '사용자_연봉 (원)': user_current_annual_salary,
            '사용자_월_실수령액 (원)': user_monthly_take_home,
            '사용자_연간_실수령액 (원)': user_annual_take_home,
            '사용자_연간_순소득 (원)': user_net_income,
            '사용자_연간_투자금액 (원)': user_investment_amount,
            '사용자_연말_자산 (원)': user_asset
        })

    df_simulation = pd.DataFrame(data)
    df_simulation.to_excel(writer, sheet_name='시뮬레이션 결과', index=False)

    # --- 시뮬레이션 시트 꾸미기 및 설정값 추가 ---
    ws_simulation = writer.sheets['시뮬레이션 결과']

    # 설정값 영역 추가 (상단에 위치)
    start_row_settings = 1
    for i, (key, value) in enumerate(settings.items()):
        ws_simulation.cell(row=start_row_settings + i, column=1).value = key
        ws_simulation.cell(row=start_row_settings + i, column=2).value = value
        ws_simulation.cell(row=start_row_settings + i, column=1).font = Font(bold=True)
        ws_simulation.cell(row=start_row_settings + i, column=1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # 연한 노란색
        ws_simulation.cell(row=start_row_settings + i, column=2).fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid") # 연한 녹색 (수정 가능)

    # 초기 자산 설정값
    ws_simulation.cell(row=start_row_settings + len(settings), column=1).value = '홍길동 시작 순자산 (원)'
    ws_simulation.cell(row=start_row_settings + len(settings), column=2).value = hong_gil_dong_initial_asset
    ws_simulation.cell(row=start_row_settings + len(settings) + 1, column=1).value = '사용자 시작 순자산 (원)'
    ws_simulation.cell(row=start_row_settings + len(settings) + 1, column=2).value = user_initial_asset
    
    # 설정값 영역 서식
    for i in range(start_row_settings, start_row_settings + len(settings) + 2):
        ws_simulation.cell(row=i, column=1).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        ws_simulation.cell(row=i, column=2).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        ws_simulation.cell(row=i, column=2).number_format = '#,##0' # 숫자 형식 적용
        
    # 결과 요약
    ws_simulation.cell(row=start_row_settings + len(settings) + 3, column=1).value = '사용자 자산 홍길동 자산 추월 시점:'
    ws_simulation.cell(row=start_row_settings + len(settings) + 3, column=2).value = f"{user_reach_hong_gil_dong_asset_year}년" if user_reach_hong_gil_dong_asset_year else "아직 미달성"
    ws_simulation.cell(row=start_row_settings + len(settings) + 3, column=1).font = Font(bold=True)
    ws_simulation.cell(row=start_row_settings + len(settings) + 3, column=2).font = Font(bold=True, color="0000FF") # 파란색

    # 데이터 시작 행 조정
    data_start_row = start_row_settings + len(settings) + 6 # 설정값 아래에 여유 공간 둔 후 데이터 시작

    # 데이터 프레임 헤더 이동
    for col_idx, col_name in enumerate(df_simulation.columns):
        ws_simulation.cell(row=data_start_row, column=col_idx + 1).value = col_name
        ws_simulation.cell(row=data_start_row, column=col_idx + 1).font = Font(bold=True)
        ws_simulation.cell(row=data_start_row, column=col_idx + 1).alignment = Alignment(horizontal='center', vertical='center')
        ws_simulation.cell(row=data_start_row, column=col_idx + 1).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # 연한 파란색

    # 데이터 쓰기
    for r_idx, row_data in enumerate(df_simulation.values):
        for c_idx, value in enumerate(row_data):
            cell = ws_simulation.cell(row=data_start_row + 1 + r_idx, column=c_idx + 1, value=value)
            if c_idx >= 1: # 숫자 형식 컬럼 (연봉, 자산 등)
                cell.number_format = '#,##0'
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    # 컬럼 너비 자동 조정
    for col_idx in range(1, len(df_simulation.columns) + 1):
        ws_simulation.column_dimensions[get_column_letter(col_idx)].width = 18

    # 틀 고정
    ws_simulation.freeze_panes = get_column_letter(1) + str(data_start_row + 1) # 데이터 헤더 아래 고정

    writer.close()
    print(f"'{file_name}' 파일이 성공적으로 생성되었습니다.")
    print("\n[중요] '실수령액 테이블' 시트에 실제 연봉별 월 실수령액 데이터를 정확히 입력해주세요.")
    print("      (연봉은 원 단위, 실수령액은 월 단위 원)")
    print("      시뮬레이션 결과는 '시뮬레이션 결과' 시트의 상단 설정값을 변경하여 조절할 수 있습니다.")


# --- 함수 실행 ---
create_financial_simulation_excel()